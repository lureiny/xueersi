from openpyxl import load_workbook
from error import ReadError, GradeError, ComprehensionError, ModelError, WeChatError
from env import PASS_LINE


class ReadExcel:
    def __init__(self, file):
        self.file = file
        self.wb = self.get_workbook()
        self.sheets = tuple(self.wb.get_sheet_names())
        self.grade1 = dict()              # 存放第一个班学生的错题和成绩
        self.grade2 = dict()              # 存放第二个班学生的错题和成绩
        self.comprehension1 = dict()      # 存放第一个班解析
        self.comprehension2 = dict()      # 存放第二个班解析
        self.wechat = dict()              # 存放学生对应的wechat
        self.model_student = None         # 存放发给学生的模板
        self.model_parents = None         # 存放发给家长的模板
        self.error_nums1 = set()          # 记录第一个班全部学生产生的错题
        self.error_nums2 = set()          # 记录第二个班全部学生产生的错题
        self.all_info = dict()            # 记录全部学生的回访信息

        self.grade_ws = None
        self.comprehension1_ws = None
        self.comprehension2_ws = None
        self.wechat_ws = None
        self.model_ws = None

        # 执行Excel文件内容检查
        self.check()
        # 检查通过后生成全部的回访信息
        self.generate_all_info()

    def check(self):
        if not self.check_sheet():
            raise ReadError("Sheet错误，请检查Sheet是否包含：grade，comprehension1，comprehension1，model，WeChat")
        else:
            self.grade_ws = self.wb.get_sheet_by_name("grade")
            self.comprehension1_ws = self.wb.get_sheet_by_name("comprehension1")
            self.comprehension2_ws = self.wb.get_sheet_by_name("comprehension2")
            self.wechat_ws = self.wb.get_sheet_by_name("WeChat")
            self.model_ws = self.wb.get_sheet_by_name("model")

        error_nums = self.check_grade()
        if error_nums:
            raise GradeError("grade表的第{}行有空值或插入的数据不合法".format("，".join(error_nums)))

        error_nums = self.check_comprehension1()
        if error_nums:
            raise ComprehensionError("comprehension1表的第{}行有空值".format("，".join(error_nums)))

        error_nums = self.check_comprehension2()
        if error_nums:
            raise ComprehensionError("comprehension2表的第{}行有空值".format("，".join(error_nums)))

        if not self.check_model():
            raise ModelError("请检查各个模板中是否包含：{姓名}, {成绩}, {错题解析}")

        error_nums = self.check_wechat()
        if error_nums:
            raise WeChatError("WeChat表的第{}行有空值".format("，".join(error_nums)))

        lack_nums = self.check_grade_comprehension1()
        if lack_nums:
            raise ComprehensionError("1班缺少以下题目的解析：\n{}".format("，".join(sorted(lack_nums))))

        lack_nums = self.check_grade_comprehension2()
        if lack_nums:
            raise ComprehensionError("2班缺少以下题目的解析：\n{}".format("，".join(sorted(lack_nums))))

        lack_student = self.check_wechat_grade1()
        if lack_student:
            raise WeChatError("缺少以下1班学生的微信联系方式：\n{}".format("，".join(lack_student)))

        lack_student = self.check_wechat_grade2()
        if lack_student:
            raise WeChatError("缺少以下2班学生的微信联系方式：\n{}".format("，".join(lack_student)))

    def get_workbook(self):
        return load_workbook(filename=self.file)

    def check_sheet(self):
        sheets = {'grade', 'comprehension1', 'model', 'WeChat', "comprehension2"}
        for sheet in self.sheets:
            sheets.add(sheet)
        return True if len(sheets) == 5 and len(self.sheets) == 5 else False

    def check_grade(self):
        if not (self.grade_ws["A1"].value == "姓名" and self.grade_ws["B1"].value == "成绩" and self.grade_ws["C1"].value == "错题" and self.grade_ws["D1"].value == "班级"):
            raise GradeError("grade表错误：请检查grade表第一行是否包含姓名、成绩、错题、班级")

        error_rows = []
        row_num = 2
        for row in self.grade_ws.iter_rows(min_row=2, max_row=self.grade_ws.max_row):
            name, grade, error_nums, class_ = row

            if not (name.value and grade.value != None and class_.value and (class_.value == 1 or class_.value == 2)):
                error_rows.append(str(row_num))
                row_num += 1
                continue

            error_nums = str(error_nums.value).split(",") if error_nums.value else []
            if class_.value == 1:
                for error_num in error_nums:
                    self.error_nums1.add(error_num)
                self.grade1[name.value] = tuple((grade.value, error_nums))
            elif class_.value == 2:
                for error_num in error_nums:
                    self.error_nums2.add(error_num)
                self.grade2[name.value] = tuple((grade.value, error_nums))
            row_num += 1
        return error_rows

    def check_model(self):
        self.model_student = tuple((self.model_ws["A1"].value, self.model_ws["B1"].value, self.model_ws["C1"].value))
        self.model_parents = tuple((self.model_ws["A2"].value, self.model_ws["B2"].value, self.model_ws["C2"].value))
        if not (self.model_student[0] and self.model_student[1] and self.model_student[2] and \
                self.model_parents[0] and self.model_parents[1] and self.model_parents[2]):
            raise ModelError("model表中缺少模版")
        if "{姓名}" in self.model_student[0] and "{成绩}" in self.model_student[0] and "{错题解析}" in self.model_student[0] and \
           "{姓名}" in self.model_student[1] and "{成绩}" in self.model_student[1] and "{错题解析}" in self.model_student[1] and \
           "{姓名}" in self.model_student[2] and "{成绩}" in self.model_student[2] and "{错题解析}" in self.model_student[2] and \
           "{姓名}" in self.model_parents[0] and "{成绩}" in self.model_parents[0] and "{错题解析}" in self.model_parents[0] and \
           "{姓名}" in self.model_parents[1] and "{成绩}" in self.model_parents[1] and "{错题解析}" in self.model_parents[1] and \
           "{姓名}" in self.model_parents[2] and "{成绩}" in self.model_parents[2] and "{错题解析}" in self.model_parents[2]:
            return True
        else:
            return False

    def check_comprehension1(self):
        if not (self.comprehension1_ws["A1"].value == "题号" and self.comprehension1_ws["B1"].value == "解析"):
            raise ComprehensionError("comprehension1表错误：请检查comprehension1表第一行是否包含题号，解析")

        error_rows = []
        row_num = 2
        for row in self.comprehension1_ws.iter_rows(min_row=2, max_row=self.comprehension1_ws.max_row):
            num, comprehension = row
            if not (num.value and comprehension.value):
                error_rows.append(str(row_num))
            self.comprehension1[str(num.value)] = str(comprehension.value)
            row_num += 1
        return error_rows

    def check_comprehension2(self):
        if not (self.comprehension2_ws["A1"].value == "题号" and self.comprehension2_ws["B1"].value == "解析"):
            raise ComprehensionError("comprehension2表错误：请检查comprehension2表第一行是否包含题号，解析")

        error_rows = []
        row_num = 2
        for row in self.comprehension2_ws.iter_rows(min_row=2, max_row=self.comprehension2_ws.max_row):
            num, comprehension = row
            if not (num.value and comprehension.value):
                error_rows.append(str(row_num))
            self.comprehension2[str(num.value)] = str(comprehension.value)
            row_num += 1
        return error_rows

    def check_wechat(self):
        if not (self.wechat_ws["A1"].value == "姓名" and self.wechat_ws["B1"].value == "学生" and self.wechat_ws["C1"].value == "家长"):
            raise WeChatError("WeChat表错误：请检查WeChat表第一行是否包含姓名，学生，家长")

        error_rows = []
        row_num = 2
        for row in self.wechat_ws.iter_rows(min_row=2, max_row=self.wechat_ws.max_row):
            name, student, parents = row
            if not (name.value and (student.value or parents.value)):                     # 判断是否有空行或者家长类型值不正确的情况
                error_rows.append(str(row_num))
            self.wechat[name.value] = tuple((self.type_change(student.value), self.type_change(parents.value)))
            row_num += 1
        return error_rows

    def check_grade_comprehension1(self):
        lack_nums = set()
        for error_num in self.error_nums1:
            if error_num not in self.comprehension1.keys():
                lack_nums.add(error_num)
        return lack_nums

    def check_grade_comprehension2(self):
        lack_nums = set()
        for error_num in self.error_nums2:
            if error_num not in self.comprehension2.keys():
                lack_nums.add(error_num)
        return lack_nums

    def check_wechat_grade1(self):
        lack_student = list()
        for student in self.grade1.keys():
            if student not in self.wechat.keys():
                lack_student.append(student)
        return lack_student

    def check_wechat_grade2(self):
        lack_student = list()
        for student in self.grade2.keys():
            if student not in self.wechat.keys():
                lack_student.append(student)
        return lack_student

    def generate_one_info(self, name, grade, error_nums, class_):
        model_index = None
        if len(name) == 3:
            temp_name = name[1:]
        else:
            temp_name = name
        comprehension = ""
        if class_ == 1:
            for num in error_nums:
                comprehension += ("{}:{}".format(num, self.comprehension1[num] + "\n"))
        else:
            for num in error_nums:
                comprehension += ("{}:{}".format(num, self.comprehension2[num] + "\n"))
        if grade >= 10:
            model_index = 0
        elif grade >= PASS_LINE:
            model_index = 1
        else:
            model_index = 2

        student = None
        parents = None

        if self.wechat[name][0]:
            student = self.model_student[model_index].format(姓名=temp_name, 成绩=grade, 错题解析=comprehension)
        if self.wechat[name][1]:
            parents = self.model_parents[model_index].format(姓名=temp_name, 成绩=grade, 错题解析=comprehension)

        return tuple((student, parents))

    def generate_all_info(self):
        for name in self.grade1:
            self.all_info[name] = self.generate_one_info(name=name, grade=self.grade1[name][0], error_nums=self.grade1[name][1], class_=1)
        for name in self.grade2:
            self.all_info[name] = self.generate_one_info(name=name, grade=self.grade2[name][0], error_nums=self.grade2[name][1], class_=2)

    @staticmethod
    def type_change(key):
        if type(key) == "int":
            return str(key)
        return key

if __name__ == '__main__':
    r = ReadExcel("test.xlsx")

    # print(r.all_info)
